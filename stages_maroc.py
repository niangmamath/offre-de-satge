# -*- coding: utf-8 -*-
"""
Veille des offres de STAGE au Maroc — TOUS domaines, TOUS types (PFE, été,
initiation, alternance).

Source principale : LinkedIn — endpoint PUBLIC "jobs-guest" (aucune connexion
requise, n'utilise pas votre compte). Mécanique de scraping/retry/cache
portée depuis sourcing-regie-banque/linkedin_sourcing_regie.py (proven en
production sur ce même type d'endpoint).
Source secondaire : Rekrute.com (cf. sources.py).

Sortie : un seul fichier Excel, un seul onglet "Stages Maroc", mis à jour de
façon INCRÉMENTALE (jamais de ligne existante écrasée — vos surlignages,
notes et suppressions manuelles sont toujours respectés).

Usage :
    python stages_maroc.py            # collecte complète (réseau)
    python stages_maroc.py reclass    # re-classe depuis le cache local, sans réseau

Réglages rapides via variables d'environnement (facultatif) :
    STG_PAGES        pages par requête LinkedIn      (défaut 2, 10 offres/page)
    STG_MAX_DETAIL    plafond de fiches détaillées     (défaut 250)
    STG_QUERY_LIMIT   limite le nb de requêtes LinkedIn (0 = toutes)
    STG_FRESH_DAYS    fenêtre de la passe fraîcheur     (défaut 7 j, 0 = off)
    STG_SIMILAR_SEEDS graines pour "offres similaires"  (défaut 20, 0 = off)

Dépendances : pip install requests beautifulsoup4 lxml openpyxl
"""
import datetime as dt
import json
import os
import random
import re
import sys
import time
from collections import Counter

# La console Windows par défaut (cp1252) plante sur "★" (U+2605) — incident
# déjà rencontré dans sourcing-regie-banque, corrigé là-bas par PYTHONUTF8=1
# en externe. Ici on le corrige dans le code : marche par défaut, sans
# dépendre de la mémoire de l'utilisateur au moment du lancement.
if sys.platform == "win32":
    for _stream in (sys.stdout, sys.stderr):
        try:
            _stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from classifier import classify_all, strip_accents, normalize_title
import sources
import db_sync

# ----------------------------------------------------------------------- CONFIG
OUTDIR = os.path.dirname(os.path.abspath(__file__))
VUES_FILE = os.path.join(OUTDIR, "annonces_vues.json")
CACHE_FILE = os.path.join(OUTDIR, "cache_stages_maroc.json")
XLSX_NAME = "Stages_Maroc.xlsx"

SEARCH_BASE = "https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search"
JOB_BASE = "https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/"

PAGES_PER_QUERY = int(os.getenv("STG_PAGES", "2"))
MAX_DETAIL_FETCH = int(os.getenv("STG_MAX_DETAIL", "250"))
QUERY_LIMIT = int(os.getenv("STG_QUERY_LIMIT", "0"))
SIMILAR_SEEDS = int(os.getenv("STG_SIMILAR_SEEDS", "20"))
FRESH_DAYS = int(os.getenv("STG_FRESH_DAYS", "7"))
SLEEP_RANGE = (1.6, 3.4)

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
]

# Grandes villes marocaines + le pays en général — pas de découpage par
# domaine métier (tous domaines confondus), contrairement à sourcing-regie-banque.
LOCATIONS = [
    "Casablanca, Maroc", "Rabat, Maroc", "Marrakech, Maroc", "Tanger, Maroc",
    "Fès, Maroc", "Agadir, Maroc", "Meknès, Maroc", "Oujda, Maroc", "Maroc",
]

# Requêtes couvrant un large éventail de domaines et d'intitulés de poste
# (demande explicite : "tous les domaines et postes de stage possibles").
# Le domaine reste un TAG informatif posé par classifier.tag_domaine, jamais
# un filtre — cette liste sert seulement à élargir ce que la recherche
# LinkedIn remonte ; une offre trouvée via "stage cybersécurité" mais tagué
# "Informatique / Data" par le classifieur reste normale.
QUERIES = [
    # --- Générique ---------------------------------------------------------
    "stage", "stage PFE", "stage de fin d'études", "stage d'été",
    "stage d'initiation", "stage alternance", "alternance", "stagiaire",
    "internship Morocco",
    # --- Informatique / Data ------------------------------------------------
    "stage développeur", "stage développement web", "stage data analyst",
    "stage data science", "stage intelligence artificielle",
    "stage cybersécurité", "stage sécurité informatique",
    "stage réseaux informatiques", "stage cloud", "stage DevOps",
    "stage UX UI design", "stage systèmes d'information",
    # --- Ingénierie / Industrie ----------------------------------------------
    "stage ingénieur", "stage génie civil", "stage génie mécanique",
    "stage génie électrique", "stage maintenance industrielle",
    "stage qualité industrielle", "stage génie industriel",
    "stage automatisme",
    # --- Finance / Comptabilité ----------------------------------------------
    "stage comptabilité", "stage audit", "stage finance",
    "stage contrôle de gestion", "stage fiscalité", "stage trésorerie",
    # --- Commerce / Marketing / Vente -----------------------------------------
    "stage marketing", "stage marketing digital", "stage commercial",
    "stage communication", "stage vente", "stage business development",
    "stage chef de produit",
    # --- RH --------------------------------------------------------------
    "stage ressources humaines", "stage recrutement", "stage paie",
    "stage formation",
    # --- Juridique ----------------------------------------------------------
    "stage juridique", "stage droit des affaires", "stage fiscaliste",
    # --- Logistique / Transport / Achats --------------------------------------
    "stage logistique", "stage supply chain", "stage achats",
    "stage transport", "stage import export",
    # --- Santé -------------------------------------------------------------
    "stage pharmacie", "stage biomédical", "stage laboratoire",
    # --- Architecture / BTP --------------------------------------------------
    "stage architecture", "stage BTP", "stage urbanisme",
    # --- Agroalimentaire / Environnement ---------------------------------------
    "stage agroalimentaire", "stage environnement", "stage agronomie",
]

WORKMODE = [
    ("hybride", "hybride"), ("télétravail", "télétravail"),
    ("teletravail", "télétravail"), ("remote", "télétravail"),
    ("à distance", "télétravail"), ("a distance", "télétravail"),
    ("full remote", "télétravail"), ("sur site", "sur site"),
    ("sur place", "sur site"), ("présentiel", "sur site"),
    ("presentiel", "sur site"), ("on-site", "sur site"),
]


# ------------------------------------------------------------------------ HTTP
# Mécanique réseau (retry/backoff/rotation UA) portée telle quelle depuis
# sourcing-regie-banque/linkedin_sourcing_regie.py — déjà éprouvée en
# production sur ce même endpoint.
def _headers():
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    }


def _sleep():
    time.sleep(random.uniform(*SLEEP_RANGE))


def _get(url, params=None, tries=3):
    for attempt in range(tries):
        try:
            r = requests.get(url, params=params, headers=_headers(), timeout=25)
        except requests.RequestException as e:
            print("      ! réseau:", e)
            time.sleep(5 * (attempt + 1))
            continue
        if r.status_code == 200:
            return r.text
        if r.status_code == 429:
            wait = 30 * (attempt + 1)
            print(f"      ! 429 (rate limit LinkedIn) -> pause {wait}s")
            time.sleep(wait)
            continue
        if r.status_code in (400, 404):
            return None
        print("      ! HTTP", r.status_code)
        time.sleep(4 * (attempt + 1))
    return None


# --------------------------------------------------------------------- PARSING
def numeric_id(card):
    if card["id"] and str(card["id"]).isdigit():
        return str(card["id"])
    m = re.search(r"(\d{6,})", card["url"])
    return m.group(1) if m else None


def search_cards(query, location, pages, recent_days=0):
    cards = []
    for p in range(pages):
        params = {"keywords": query, "location": location, "start": p * 10}
        if recent_days:
            params["f_TPR"] = f"r{int(recent_days) * 86400}"
        txt = _get(SEARCH_BASE, params)
        _sleep()
        if not txt:
            break
        soup = BeautifulSoup(txt, "lxml")
        lis = soup.select("li")
        if not lis:
            break
        got = 0
        for li in lis:
            base = li.select_one("div.base-card") or li
            urn = base.get("data-entity-urn", "") if base else ""
            jid = urn.split(":")[-1] if urn else ""
            a = (li.select_one("a.base-card__full-link")
                 or li.select_one("a[href*='/jobs/view/']"))
            link = a["href"].split("?")[0] if a and a.has_attr("href") else ""
            title = li.select_one("h3.base-search-card__title")
            comp = li.select_one("h4.base-search-card__subtitle")
            loc = li.select_one("span.job-search-card__location")
            tnode = li.select_one("time")
            posted = ""
            if tnode:
                posted = tnode.get("datetime") or tnode.get_text(strip=True)
            if not (title and link):
                continue
            cards.append({
                "id": jid or link,
                "poste": title.get_text(strip=True),
                "entite": comp.get_text(strip=True) if comp else "",
                "ville": loc.get_text(strip=True) if loc else "",
                "url": link,
                "publication": posted,
            })
            got += 1
        if got == 0:
            break
    return cards


def similar_cards(job_url):
    """Offres SIMILAIRES suggérées par LinkedIn en bas d'une fiche — souvent
    plus fraîches que le résultat de recherche."""
    if not job_url:
        return []
    txt = _get(job_url.split("?")[0])
    _sleep()
    if not txt:
        return []
    soup = BeautifulSoup(txt, "lxml")
    cards, vus = [], set()
    for base in soup.select("div.base-card"):
        a = base.select_one("a[href*='/jobs/view/']")
        if not a or not a.has_attr("href"):
            continue
        link = a["href"].split("?")[0]
        m = re.search(r"-(\d{8,})", link)
        jid = m.group(1) if m else link
        if jid in vus:
            continue
        vus.add(jid)
        title = base.select_one("h3")
        comp = base.select_one("h4")
        loc = base.select_one("[class*=location]")
        tnode = base.select_one("time")
        if not title:
            continue
        cards.append({
            "id": jid,
            "poste": title.get_text(strip=True),
            "entite": comp.get_text(strip=True) if comp else "",
            "ville": loc.get_text(strip=True) if loc else "",
            "url": link,
            "publication": (tnode.get("datetime") or tnode.get_text(strip=True)) if tnode else "",
        })
    return cards


def fetch_detail(jid):
    txt = _get(JOB_BASE + str(jid))
    _sleep()
    if not txt:
        return {}
    soup = BeautifulSoup(txt, "lxml")
    node = (soup.select_one("div.show-more-less-html__markup")
            or soup.select_one("div.description__text"))
    desc = node.get_text(" ", strip=True) if node else ""
    crit = {}
    for li in soup.select("ul.description__job-criteria-list li"):
        lab = li.select_one("h3")
        val = li.select_one("span")
        if lab and val:
            crit[lab.get_text(strip=True).lower()] = val.get_text(strip=True)

    na = (soup.select_one("figcaption.num-applicants__caption")
          or soup.select_one("span.num-applicants__caption")
          or soup.select_one("[class*='num-applicants__caption']"))
    nb_candidats_txt = na.get_text(strip=True) if na else ""
    pt = soup.select_one("span.posted-time-ago__text")
    posted_relative = pt.get_text(strip=True) if pt else ""
    low = txt.lower()
    cloturee = ("n'accepte plus" in low or "no longer accepting" in low
                or "plus de candidatures" in low)
    republication = ""
    if "republi" in low or "reposted" in low:
        i = low.find("republi") if "republi" in low else low.find("reposted")
        republication = re.sub(r"\s+", " ", txt[i:i + 40]).strip()

    return {"desc": desc, "crit": crit, "nb_candidats_txt": nb_candidats_txt,
            "posted_relative": posted_relative, "cloturee": cloturee,
            "republication": republication}


def work_mode(text):
    t = text.lower()
    for k, v in WORKMODE:
        if k in t:
            return v
    return "NC"


def duration(text):
    t = text.lower()
    rng = re.search(r"\d{1,2}\s*(?:[-/à]|a)\s*\d{1,2}\s*mois", t)
    if rng:
        return rng.group(0)
    one = re.search(r"\d{1,2}\s*mois", t)
    if one:
        return one.group(0)
    if "longue durée" in t or "longue duree" in t:
        return "longue durée"
    return "NC"


def mission_snippet(desc):
    if not desc:
        return "NC"
    parts = re.split(r"(?<=[\.\!\?])\s", desc.strip())
    snip = (parts[0] if parts else desc).strip()
    return (snip[:160] + "…") if len(snip) > 160 else (snip or "NC")


def _entite_depuis_url(url):
    m = re.search(r"-at-([a-z0-9-]+?)-\d{6,}", (url or "").lower())
    if not m:
        return ""
    return m.group(1).replace("-", " ").strip()


def to_annonce(card, detail):
    iso = str(card.get("publication") or "")
    iso = iso[:10] if re.match(r"\d{4}-\d{2}-\d{2}", iso) else ""
    entite = card.get("entite", "") or _entite_depuis_url(card.get("url", ""))
    emp = ""
    for lab, val in (detail.get("crit") or {}).items():
        if "type" in lab and "emploi" in lab:
            emp = val
    desc = detail.get("desc", "")
    return {
        "poste": card.get("poste", ""),
        "entite": entite,
        "ville": card.get("ville", ""),
        "url": (card.get("url", "") or "").split("?")[0],
        "date_pub": iso,
        "posted_relative": detail.get("posted_relative", ""),
        "texte": desc,
        "emploi_label": emp,
        "nb_candidats_txt": detail.get("nb_candidats_txt", ""),
        "cloturee": detail.get("cloturee", False),
        "republication": detail.get("republication", ""),
        "open_confirme": False,
        "source": "LinkedIn",
        "mission": mission_snippet(desc),
        "lieu": work_mode(desc),
        "duree": duration(desc),
    }


# Titre/entreprise : au moins un signal "stage" pour valoir la peine d'être
# téléchargé en détail — filtre grossier de pré-tri, la vraie décision est
# prise par classifier.detect_stage() sur le texte complet.
_PRE_KW = ["stage", "stagiaire", "pfe", "alternance", "alternant",
           "intern", "apprenti"]


def _pre_filter(card):
    blob = f" {card.get('poste', '')} ".lower()
    return any(k in blob for k in _PRE_KW)


def harvest():
    """Collecte multi-sources -> liste d'ANNONCES, mise en cache local
    (permet le mode `reclass`, hors ligne)."""
    print("\n### Collecte — stages Maroc ###")

    known = {}
    try:
        with open(CACHE_FILE, encoding="utf-8") as f:
            for a in json.load(f):
                u = (a.get("url") or "").split("?")[0].rstrip("/")
                if u and a.get("source") == "LinkedIn":
                    known[u] = a
    except Exception:
        pass

    queries = QUERIES[:QUERY_LIMIT] if QUERY_LIMIT else QUERIES

    # 1) LinkedIn
    raw = {}
    for loc in LOCATIONS:
        for q in queries:
            print(f"  · LinkedIn '{q}' @ {loc}")
            for c in search_cards(q, loc, PAGES_PER_QUERY):
                raw.setdefault(c["id"], c)
    print(f"  {len(raw)} offres LinkedIn (avant fiche détaillée).")

    if FRESH_DAYS:
        print(f"  · Passe fraîcheur (<= {FRESH_DAYS} j)...")
        neuf = 0
        for loc in LOCATIONS:
            for q in queries:
                for c in search_cards(q, loc, 1, recent_days=FRESH_DAYS):
                    if c["id"] not in raw:
                        raw[c["id"]] = c
                        neuf += 1
        print(f"    +{neuf} offres récentes que la recherche normale ratait.")

    if SIMILAR_SEEDS:
        graines = [c for c in raw.values() if _pre_filter(c)][:SIMILAR_SEEDS]
        print(f"  · Offres similaires LinkedIn (à partir de {len(graines)} offres)...")
        trouve = 0
        for c in graines:
            for s in similar_cards(c.get("url", "")):
                if s["id"] not in raw:
                    raw[s["id"]] = s
                    trouve += 1
        print(f"    +{trouve} offres suggérées par LinkedIn.")

    annonces = []
    fetched = reused = 0
    for card in raw.values():
        if not _pre_filter(card):
            continue
        u = (card.get("url") or "").split("?")[0].rstrip("/")
        if u in known:
            annonces.append(known[u])
            reused += 1
            continue
        detail = {}
        nid = numeric_id(card)
        if nid and fetched < MAX_DETAIL_FETCH:
            detail = fetch_detail(nid)
            fetched += 1
        annonces.append(to_annonce(card, detail))
    print(f"  LinkedIn : {len(annonces)} annonces "
          f"({fetched} fiches téléchargées, {reused} réutilisées du cache).")

    # 2) Sources non-LinkedIn (Rekrute.com, etc. — cf. sources.py)
    print("  · Sources complémentaires...")
    autres = sources.collect_all()
    annonces += autres
    print(f"  TOTAL {len(annonces)} annonces (LinkedIn + {len(autres)} autres sources).")

    # Fusion avec le cache précédent : une offre disparue des résultats reste
    # connue (sinon elle devient une ligne orpheline jamais re-filtrée).
    fusion = {}
    try:
        with open(CACHE_FILE, encoding="utf-8") as f:
            for a in json.load(f):
                u = (a.get("url") or "").split("?")[0].rstrip("/")
                if u:
                    fusion[u] = a
    except Exception:
        pass
    for a in annonces:
        u = (a.get("url") or "").split("?")[0].rstrip("/")
        if u:
            fusion[u] = a
    tout = list(fusion.values())
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(tout, f, ensure_ascii=False)
    print(f"  Cache : {len(tout)} annonces connues ({len(annonces)} vues ce run).")
    return tout


def process(annonces, vues, today_iso):
    """annonces -> classification complète -> fusion avec l'historique."""
    items = classify_all(annonces, today_iso)

    now_disp = dt.datetime.now().strftime("%d/%m/%Y %H:%M")
    nouveaux = []
    for a in items:
        url = a["url"]
        rec = vues.get(url)
        if rec:
            a["premiere_detection"] = rec.get("premiere_detection", now_disp)
            a["nouveau"] = False
        else:
            a["premiere_detection"] = now_disp
            a["nouveau"] = True
            if a["verdict"] != "ÉCARTÉE":
                nouveaux.append(a)
        a["derniere_verification"] = now_disp
        vues[url] = {
            "premiere_detection": a["premiere_detection"],
            "derniere_verification": now_disp,
            "verdict": a["verdict"],
            "poste": a["poste"],
            "entite": a["entite"],
        }

    conv = [a for a in items if a["verdict"] != "ÉCARTÉE"]
    c = Counter(a["fenetre"] for a in conv)
    print(f"\n  {len(items)} annonces classées : {len(conv)} CONVENABLES "
          f"(sur {len(items)}), dont {len(nouveaux)} nouvelles.")
    for f in ["NOUVEAU", "OUVERTE", "ROUVERTE", "VIVIER", "AGÉE", "INCONNUE"]:
        if c.get(f):
            print(f"       {c[f]:>3}  {f}")
    for a in nouveaux:
        print(f"  ★ NOUVEAU | {a['fenetre']} | {a['type_stage']} | {a['entite']} | {a['poste']}")
    return items


# -------------------------------------------------------------------- EXCEL
def fmt_pub(p):
    if not p:
        return "NC"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", p)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    return p


def _pub_display(a):
    iso = a.get("date_pub_iso")
    if iso:
        return fmt_pub(iso)
    return a.get("posted_relative") or "NC"


SHEET_TITLE = "Stages Maroc"

HEADERS = ["", "poste", "entreprise", "ville", "description", "domaine",
           "type de stage", "durée", "SOURCE", "publication",
           "candidats", "âge (j)", "fenêtre", "1re détection",
           "dern. vérif.", "provenance"]
WIDTHS = [12, 34, 24, 15, 46, 20, 16, 11, 14, 12, 10, 8, 12, 16, 16, 16]

FENETRE_FILL = {
    "NOUVEAU": "C6EFCE",
    "OUVERTE": "E2EFDA",
    "ROUVERTE": "DDEBF7",
    "VIVIER": "FFF2CC",
    "AGÉE": "E7E6E6",
    "INCONNUE": "FFFFFF",
}
NEW_ROW_FILL = "D9EAD3"
NEW_ROW_FONT = Font(bold=True, color="1F4E78")


def _dedup_key(a):
    return (strip_accents(a.get("entite", "")).lower().strip(),
            normalize_title(a.get("poste", "")))


def is_convenable(a):
    return a["verdict"] != "ÉCARTÉE"


def _retirer_existant(a):
    """Ligne DÉJÀ dans le fichier à retirer — uniquement pour une raison
    DURE (junk), jamais pour l'âge : une offre de stage périmée reste une
    trace utile, pas une erreur de classification."""
    return bool(not a.get("est_stage") or a.get("hors_maroc_flag"))


def _write_sheet(wb, rows, today):
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="0563C1", underline="single")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet(SHEET_TITLE)
    ws["A1"] = f"MISE A JOUR : {today}"
    ws["A1"].font = Font(bold=True, color="C00000")
    ws["A1"].comment = Comment(
        "Fenêtre : NOUVEAU (<=7j) · OUVERTE (<=21j) · ROUVERTE (republiée) · "
        "VIVIER (encore listée, source à annonces actives) · AGÉE (pas de "
        "signal de fraîcheur) · INCONNUE (date introuvable).\n"
        "Poste en gras + ★ = nouveau depuis le dernier passage.\n"
        "Domaine et type de stage sont INFORMATIFS (aucune offre n'est "
        "écartée pour son domaine ou son type de stage).",
        "scraper")
    for ci, h in enumerate(HEADERS[1:], start=2):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    r = 2
    src_col = HEADERS.index("SOURCE") + 1
    for a in rows:
        is_new = bool(a.get("nouveau"))
        fill_hex = FENETRE_FILL.get(a["fenetre"], "FFFFFF")
        if is_new:
            fill_hex = NEW_ROW_FILL
        fill = PatternFill("solid", fgColor=fill_hex)
        nb = a.get("nb_candidats_int")
        age = a.get("age_jours")
        source = a.get("source", "LinkedIn")
        values = [
            None, a["poste"], a["entite"], (a.get("ville") or "").upper(),
            a.get("mission", "NC"), a.get("domaine", "Autre"),
            a.get("type_stage", "Non précisé"), a.get("duree", "NC"),
            source, _pub_display(a),
            nb if nb is not None else "NC", age if age is not None else "NC",
            a["fenetre"], a.get("premiere_detection", today),
            a.get("derniere_verification", today), source,
        ]
        for ci, v in enumerate(values, start=1):
            cell = ws.cell(r, ci, v)
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 5))
        if is_new:
            ws.cell(r, 2).font = NEW_ROW_FONT
            ws.cell(r, 2).value = f"★ {ws.cell(r, 2).value}"
            ws.cell(r, 2).fill = fill
        src = ws.cell(r, src_col)
        src.hyperlink = a["url"]
        src.font = link_font
        ws.row_dimensions[r].height = 28
        r += 1

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"B1:{get_column_letter(len(HEADERS))}{max(1, r - 1)}"
    return ws


def write_excel(path, items, today):
    """Création complète (1re exécution ou fichier illisible)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    conv, seen_k = [], set()
    for a in items:
        if not is_convenable(a):
            continue
        k = _dedup_key(a)
        if k in seen_k:
            continue
        seen_k.add(k)
        conv.append(a)
    _write_sheet(wb, conv, today)
    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = path.replace(".xlsx", "_NEW.xlsx")
        wb.save(alt)
        print(f"  ! '{os.path.basename(path)}' est ouvert dans Excel — "
              f"écrit dans '{os.path.basename(alt)}'.")
        return alt


# --------------------------------------------------- MISE A JOUR INCREMENTALE
def _colkey(s):
    s = (s or "").lower().replace("œ", "oe").replace("(j)", "")
    return re.sub(r"[^a-z0-9]", "", strip_accents(s))


def _row_values_map(a, today):
    nb = a.get("nb_candidats_int")
    age = a.get("age_jours")
    source = a.get("source", "LinkedIn")
    return {
        "poste": a["poste"], "entreprise": a["entite"],
        "ville": (a.get("ville") or "").upper(),
        "description": a.get("mission", "NC"),
        "domaine": a.get("domaine", "Autre"),
        "typedestage": a.get("type_stage", "Non précisé"),
        "duree": a.get("duree", "NC"), "source": source,
        "publication": _pub_display(a),
        "candidats": nb if nb is not None else "NC",
        "age": age if age is not None else "NC",
        "fenetre": a["fenetre"],
        "1redetection": a.get("premiere_detection", today),
        "dernverif": a.get("derniere_verification", today),
        "provenance": source,
    }


def _header_map(ws):
    m = {}
    for c in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(1, c).value
        if v:
            m[_colkey(str(v))] = c
    return m


def _existing_urls(ws, src_col):
    urls = set()
    for r in range(2, (ws.max_row or 1) + 1):
        cell = ws.cell(r, src_col)
        if cell.hyperlink and cell.hyperlink.target:
            urls.add(cell.hyperlink.target.split("?")[0].rstrip("/"))
    return urls


def _delete_rows_safe(ws, a_retirer, src_col):
    """Supprime des lignes sans casser les hyperliens (piège openpyxl :
    delete_rows() ne déplace pas les hyperliens stockés par adresse de
    cellule)."""
    if not a_retirer:
        return
    a_retirer = sorted(set(a_retirer))
    liens = {}
    for rr in range(2, (ws.max_row or 1) + 1):
        c = ws.cell(rr, src_col)
        if c.hyperlink:
            liens[rr] = c.hyperlink.target
            c.hyperlink = None
    for rr in reversed(a_retirer):
        ws.delete_rows(rr)
    for old_r, target in liens.items():
        if old_r in a_retirer:
            continue
        decalage = sum(1 for s in a_retirer if s < old_r)
        ws.cell(old_r - decalage, src_col).hyperlink = target


def _last_data_row(ws, poste_col):
    last = 1
    for r in range(2, (ws.max_row or 1) + 1):
        if ws.cell(r, poste_col).value not in (None, ""):
            last = r
    return last


def _append_offer(ws, r, a, hmap, today):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor=NEW_ROW_FILL)
    link_font = Font(color="0563C1", underline="single")
    vals = _row_values_map(a, today)
    for key, col in hmap.items():
        if key not in vals:
            continue
        v = vals[key]
        if key == "poste":
            v = f"★ {v}"
        cell = ws.cell(r, col, v)
        cell.border = border
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=(key == "description"))
        if key == "source":
            cell.hyperlink = a["url"]
            cell.font = link_font
        elif key == "poste":
            cell.font = NEW_ROW_FONT
    ws.row_dimensions[r].height = 28


def update_excel(path, items, today):
    """Met à jour le fichier SANS rien écraser : les modifications manuelles
    (remplissage, surlignage, suppressions) sont conservées. Seules les
    nouvelles offres convenables sont ajoutées, marquées ★. Si le fichier
    n'existe pas encore -> création complète."""
    if not os.path.exists(path):
        return write_excel(path, items, today)
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"  ! fichier illisible ({e}) -> recréation complète.")
        return write_excel(path, items, today)

    if SHEET_TITLE not in wb.sheetnames:
        _write_sheet(wb, [], today)  # feuille vide, ajout géré ci-dessous
    ws = wb[SHEET_TITLE]
    hmap = _header_map(ws)
    src_col, poste_col = hmap.get("source"), hmap.get("poste")
    if not src_col or not poste_col:
        print("  ! en-tête introuvable — abandon (fichier non reconnu).")
        return write_excel(path, items, today)

    # 1) Nettoyage : retire les lignes devenues junk (plus un stage / hors
    #    Maroc) parmi celles qu'on peut encore identifier par URL.
    juges = {}
    for a in items:
        u = (a.get("url") or "").split("?")[0].rstrip("/")
        if u:
            juges[u] = a
    a_retirer = []
    for rr in range(2, (ws.max_row or 1) + 1):
        if not ws.cell(rr, poste_col).value:
            continue
        c = ws.cell(rr, src_col)
        a = juges.get(c.hyperlink.target.split("?")[0].rstrip("/")) if c.hyperlink else None
        if a is not None and _retirer_existant(a):
            a_retirer.append(rr)
    _delete_rows_safe(ws, a_retirer, src_col)
    if a_retirer:
        print(f"  {len(a_retirer)} ligne(s) devenue(s) non convenable(s) retirée(s).")

    # 1bis) Efface les ★ de la mise à jour précédente (★ = "nouveau à CE run").
    deetoilees = 0
    for rr in range(2, (ws.max_row or 1) + 1):
        v = ws.cell(rr, poste_col).value
        if isinstance(v, str) and v.lstrip().startswith("★"):
            ws.cell(rr, poste_col).value = v.lstrip().lstrip("★").lstrip()
            deetoilees += 1
    if deetoilees:
        print(f"  {deetoilees} étoile(s) de la MAJ précédente effacée(s).")

    existing = _existing_urls(ws, src_col)
    r = _last_data_row(ws, poste_col)

    conv, seen_k = [], set()
    for a in items:
        if not is_convenable(a):
            continue
        k = _dedup_key(a)
        if k in seen_k:
            continue
        seen_k.add(k)
        conv.append(a)

    verif_col = hmap.get("dernverif")
    vus = {(a.get("url") or "").split("?")[0].rstrip("/") for a in conv}
    if verif_col:
        for rr in range(2, r + 1):
            c = ws.cell(rr, src_col)
            if c.hyperlink and c.hyperlink.target.split("?")[0].rstrip("/") in vus:
                ws.cell(rr, verif_col).value = today

    added = 0
    for a in conv:
        u = (a.get("url") or "").split("?")[0].rstrip("/")
        if u in existing or not a.get("nouveau"):
            continue
        r += 1
        _append_offer(ws, r, a, hmap, today)
        existing.add(u)
        added += 1
    ws.cell(1, 1).value = f"MISE A JOUR : {today}"
    print(f"  {added} nouvelle(s) offre(s) ajoutée(s) à la suite.")

    try:
        wb.save(path)
        print(f"  => {added} nouvelle(s) offre(s) (vos modifications conservées).")
        return path
    except PermissionError:
        alt = path.replace(".xlsx", "_NEW.xlsx")
        wb.save(alt)
        print(f"  ! '{os.path.basename(path)}' est ouvert dans Excel — "
              f"écrit dans '{os.path.basename(alt)}'.")
        return alt


# ------------------------------------------------------------------- HISTORIQUE
def load_vues():
    try:
        with open(VUES_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_vues(vues):
    with open(VUES_FILE, "w", encoding="utf-8") as f:
        json.dump(vues, f, ensure_ascii=False, indent=1)


# -------------------------------------------------------------------- MAIN
def main():
    args = [a.lower() for a in sys.argv[1:]]
    reclass = bool(args) and args[0] in ("reclass", "recla", "cache", "retune")

    today_iso = dt.date.today().isoformat()
    today_disp = dt.datetime.now().strftime("%d/%m/%Y %H:%M")
    vues = load_vues()

    if reclass:
        try:
            with open(CACHE_FILE, encoding="utf-8") as f:
                annonces = json.load(f)
            print("\n### Reclassement (cache, hors ligne) ###")
        except FileNotFoundError:
            print("  ! pas de cache local. Lancez d'abord une collecte complète.")
            annonces = []
    else:
        annonces = harvest()

    items = process(annonces, vues, today_iso)

    path = os.path.join(OUTDIR, os.getenv("STG_OUT", XLSX_NAME))
    written = update_excel(path, items, today_disp)
    db_sync.sync(items, is_convenable)

    if os.getenv("STG_OUT"):
        print("  ! Run de relecture (STG_OUT) : historique NON enregistré.")
    elif os.path.abspath(written) == os.path.abspath(path):
        save_vues(vues)
    else:
        print("  ! Historique NON enregistré (fichier verrouillé) : ces "
              "offres restent 'nouvelles' et reviendront au prochain run.")

    conv = sum(1 for a in items if is_convenable(a))
    print(f"\n===> {conv} offres de stage convenables dans :\n     {written}")


if __name__ == "__main__":
    main()
